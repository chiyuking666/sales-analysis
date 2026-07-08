import pandas as pd
import os
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

def add_pie_chart(sheet, title: str, anchor: str) -> None:
    if sheet.max_row <= 1:
        return
    labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
    chart = PieChart()
    chart.title = title
    chart.add_data(data, titles_from_data = True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    sheet.add_chart(chart, anchor)


def export_sales(result_dict: dict,path: str) -> str:
    os.makedirs(path, exist_ok=True)
    path = os.path.join(path,'summary.xlsx')
    chart_config = {
        '地区销售统计': ('地区销售占比', 'D2'),
        '产品销售统计': ('产品销售占比', 'D2'),
        }
    try:
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            for name,df in result_dict.items():
                df.to_excel(writer,sheet_name=name,index=False)
            for sheet_name, (title, anchor) in chart_config.items():
                sheet = writer.sheets.get(sheet_name)
                add_pie_chart(sheet, title, anchor)
    except PermissionError:
        raise PermissionError(
            f'无法写入 {path}，请先关闭已打开的Excel文件。'
        )
    return path
